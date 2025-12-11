# app/file_access/document_ops/excel_ops.py
"""
Protocol-agnostic Excel operations.

All functions take a FileStorageProvider and operate on bytes,
making them work with any storage backend (local, SMB, NFS, etc.).
"""
import io
from typing import Any, Dict, List, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import structlog

from app.file_access.base_fs import FileStorageProvider

logger = structlog.get_logger()


async def read_excel(
    provider: FileStorageProvider,
    path: str,
    sheet_name: Optional[str] = None,
    read_only: bool = True
) -> Workbook:
    """
    Read Excel file from any provider.
    
    Args:
        provider: FileStorageProvider instance (connected)
        path: Path to Excel file
        sheet_name: Specific sheet to activate (optional)
        read_only: Open in read-only mode for performance
        
    Returns:
        openpyxl Workbook object
        
    Raises:
        FileNotFoundError: If file doesn't exist
    """
    try:
        logger.debug("reading_excel", path=path, sheet_name=sheet_name)
        
        # Read file bytes from provider
        data = await provider.read_file(path)
        
        # Load workbook from bytes
        buffer = io.BytesIO(data)
        wb = load_workbook(buffer, read_only=read_only, data_only=True)
        
        # Activate specific sheet if requested
        if sheet_name and sheet_name in wb.sheetnames:
            wb.active = wb[sheet_name]
        
        logger.info("excel_read_success", path=path, sheets=wb.sheetnames)
        return wb
        
    except Exception as e:
        logger.error("excel_read_failed", path=path, error=str(e))
        raise


async def write_excel(
    provider: FileStorageProvider,
    path: str,
    workbook: Workbook,
    overwrite: bool = False
) -> bool:
    """
    Write Excel workbook to any provider.
    
    Args:
        provider: FileStorageProvider instance (connected)
        path: Destination path
        workbook: openpyxl Workbook to write
        overwrite: Whether to overwrite existing file
        
    Returns:
        True if successful
        
    Raises:
        FileExistsError: If file exists and overwrite=False
    """
    try:
        logger.debug("writing_excel", path=path, overwrite=overwrite)
        
        # Save workbook to bytes
        buffer = io.BytesIO()
        workbook.save(buffer)
        data = buffer.getvalue()
        
        # Write bytes to provider
        result = await provider.write_file(path, data, overwrite=overwrite)
        
        if result.success:
            logger.info("excel_write_success", path=path, size=len(data))
            return True
        else:
            logger.error("excel_write_failed", path=path, error=result.error)
            raise Exception(f"Failed to write Excel: {result.error}")
            
    except Exception as e:
        logger.error("excel_write_failed", path=path, error=str(e))
        raise


async def update_excel(
    provider: FileStorageProvider,
    path: str,
    updates: Dict[str, Dict[str, Any]],
    sheet_name: str = "Sheet1"
) -> bool:
    """
    Update specific cells in Excel file.
    
    Args:
        provider: FileStorageProvider instance (connected)
        path: Path to Excel file
        updates: Dict of {cell_address: value}, e.g., {"A1": "Name", "B2": 100}
        sheet_name: Sheet to update (default: "Sheet1")
        
    Returns:
        True if successful
        
    Example:
        updates = {
            "A1": "Customer Name",
            "B1": "ACME Corp",
            "A2": "Amount",
            "B2": 1250.50
        }
        await update_excel(provider, "/quotes/Q001.xlsx", updates)
    """
    try:
        logger.debug("updating_excel", path=path, updates_count=len(updates))
        
        # Read existing workbook
        wb = await read_excel(provider, path, read_only=False)
        
        # Get or create sheet
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]
        
        # Apply updates
        for cell_address, value in updates.items():
            ws[cell_address] = value
        
        # Write back
        success = await write_excel(provider, path, wb, overwrite=True)
        
        logger.info("excel_update_success", path=path, updates=list(updates.keys()))
        return success
        
    except Exception as e:
        logger.error("excel_update_failed", path=path, error=str(e))
        raise


async def create_excel_from_data(
    provider: FileStorageProvider,
    path: str,
    data: List[List[Any]],
    sheet_name: str = "Sheet1",
    headers: Optional[List[str]] = None,
    overwrite: bool = False
) -> bool:
    """
    Create Excel file from tabular data.
    
    Args:
        provider: FileStorageProvider instance (connected)
        path: Destination path
        data: List of rows (each row is a list of values)
        sheet_name: Sheet name
        headers: Column headers (optional, prepended to data)
        overwrite: Whether to overwrite existing file
        
    Returns:
        True if successful
        
    Example:
        data = [
            ["Customer A", 1000.00, "2024-01-01"],
            ["Customer B", 1500.50, "2024-01-02"],
        ]
        headers = ["Customer", "Amount", "Date"]
        
        await create_excel_from_data(
            provider,
            "/reports/sales.xlsx",
            data,
            headers=headers
        )
    """
    try:
        logger.debug("creating_excel", path=path, rows=len(data))
        
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Add headers if provided
        if headers:
            ws.append(headers)
        
        # Add data rows
        for row in data:
            ws.append(row)
        
        # Write to provider
        success = await write_excel(provider, path, wb, overwrite=overwrite)
        
        logger.info("excel_create_success", path=path, rows=len(data))
        return success
        
    except Exception as e:
        logger.error("excel_create_failed", path=path, error=str(e))
        raise


async def read_excel_as_dict(
    provider: FileStorageProvider,
    path: str,
    sheet_name: Optional[str] = None,
    header_row: int = 1
) -> List[Dict[str, Any]]:
    """
    Read Excel as list of dictionaries (header row as keys).
    
    Args:
        provider: FileStorageProvider instance (connected)
        path: Path to Excel file
        sheet_name: Sheet to read (optional, uses active sheet)
        header_row: Row number for headers (1-indexed, default: 1)
        
    Returns:
        List of dicts, one per data row
        
    Example:
        # Excel:
        # | Name     | Age | City     |
        # | John Doe | 30  | New York |
        # | Jane Doe | 25  | Boston   |
        
        result = await read_excel_as_dict(provider, "/data.xlsx")
        # [
        #   {"Name": "John Doe", "Age": 30, "City": "New York"},
        #   {"Name": "Jane Doe", "Age": 25, "City": "Boston"}
        # ]
    """
    try:
        logger.debug("reading_excel_as_dict", path=path, sheet_name=sheet_name)
        
        wb = await read_excel(provider, path, sheet_name=sheet_name)
        ws = wb.active
        
        # Get headers from specified row
        headers = []
        for cell in ws[header_row]:
            headers.append(cell.value)
        
        # Read data rows
        result = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            row_dict = {}
            for header, value in zip(headers, row):
                if header:  # Skip empty headers
                    row_dict[header] = value
            if any(row_dict.values()):  # Skip empty rows
                result.append(row_dict)
        
        logger.info("excel_read_as_dict_success", path=path, rows=len(result))
        return result
        
    except Exception as e:
        logger.error("excel_read_as_dict_failed", path=path, error=str(e))
        raise


async def append_excel_row(
    provider: FileStorageProvider,
    path: str,
    row_data: List[Any],
    sheet_name: str = "Sheet1"
) -> bool:
    """
    Append row to existing Excel file.
    
    Args:
        provider: FileStorageProvider instance (connected)
        path: Path to Excel file
        row_data: List of values to append
        sheet_name: Sheet to append to
        
    Returns:
        True if successful
        
    Example:
        await append_excel_row(
            provider,
            "/quotes/log.xlsx",
            ["2024-01-15", "Q001", "ACME Corp", 1000.00]
        )
    """
    try:
        logger.debug("appending_excel_row", path=path)
        
        # Read existing workbook
        wb = await read_excel(provider, path, read_only=False)
        
        # Get or create sheet
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]
        
        # Append row
        ws.append(row_data)
        
        # Write back
        success = await write_excel(provider, path, wb, overwrite=True)
        
        logger.info("excel_append_success", path=path)
        return success
        
    except Exception as e:
        logger.error("excel_append_failed", path=path, error=str(e))
        raise
