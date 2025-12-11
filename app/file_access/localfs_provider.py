# app/file_access/localfs_provider.py
"""
Local filesystem provider for FAAL.

Supports any NAS mounted at a local path.
Generic implementation - not tied to any specific NAS brand.
"""
import os
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
import aiofiles
import aiofiles.os
from filelock import FileLock, Timeout
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.file_access.base import FileStorageProvider, FileOperationResult, FileMetadata
from app.monitoring.logger import log


class LocalFSProvider(FileStorageProvider):
    """
    Local filesystem provider.
    
    Works with:
    - Local directories
    - Network-mounted NAS (NFS, SMB/CIFS, etc.)
    - Any path accessible via standard filesystem operations
    
    Config schema:
    {
        "base_path": "/path/to/storage",  # Required
        "excel_filename": "preventivi.xlsx",  # Optional, default: "preventivi.xlsx"
        "lock_timeout": 30,  # Optional, seconds to wait for file lock
        "create_dirs": true  # Optional, create directories if missing
    }
    """
    
    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        
        # Validate required config
        if "base_path" not in config:
            raise ValueError("LocalFSProvider requires 'base_path' in config")
        
        self.base_path = Path(config["base_path"])
        self.excel_filename = config.get("excel_filename", "preventivi.xlsx")
        self.lock_timeout = config.get("lock_timeout", 30)
        self.create_dirs = config.get("create_dirs", True)
        
        log("INFO", f"LocalFSProvider initialized with base_path={self.base_path}", 
            module="localfs_provider")
    
    def _resolve_path(self, path: str) -> Path:
        """Resolve relative path to absolute path within base_path."""
        resolved = (self.base_path / path).resolve()
        
        # Security check: ensure resolved path is within base_path
        try:
            resolved.relative_to(self.base_path)
        except ValueError:
            raise PermissionError(f"Access denied: path '{path}' is outside base_path")
        
        return resolved
    
    async def update_quote_excel(
        self,
        tenant_id: str,
        quote: Any,
        customer: Dict[str, Any]
    ) -> FileOperationResult:
        """
        Update Excel spreadsheet with quote data.
        
        Uses FileLock to prevent concurrent write conflicts.
        Creates file if it doesn't exist.
        """
        excel_path = self._resolve_path(self.excel_filename)
        lock_path = excel_path.with_suffix(".lock")
        
        try:
            # Ensure directory exists
            if self.create_dirs:
                excel_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Acquire file lock
            lock = FileLock(str(lock_path), timeout=self.lock_timeout)
            
            with lock:
                # Load or create workbook
                if excel_path.exists():
                    wb = load_workbook(str(excel_path))
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Preventivi"
                    
                    # Create header row
                    headers = [
                        "ID", "Data", "Cliente", "Email", "Telefono",
                        "Totale", "Stato", "Note"
                    ]
                    ws.append(headers)
                    
                    # Style header
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", 
                                             fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Check if quote already exists (by ID in first column)
                quote_id_str = str(quote.id)
                row_index = None
                
                for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=False), start=2):
                    if row[0].value and str(row[0].value) == quote_id_str:
                        row_index = idx
                        break
                
                # Prepare row data
                row_data = [
                    str(quote.id),
                    quote.created_at.strftime("%Y-%m-%d %H:%M") if quote.created_at else "",
                    customer.get("name", ""),
                    customer.get("email", ""),
                    customer.get("phone", ""),
                    float(getattr(quote, "total", 0.0)),
                    getattr(quote, "status", "pending"),
                    getattr(quote, "notes", "")
                ]
                
                if row_index:
                    # Update existing row
                    for col_idx, value in enumerate(row_data, start=1):
                        ws.cell(row=row_index, column=col_idx, value=value)
                    operation = "updated"
                else:
                    # Append new row
                    ws.append(row_data)
                    operation = "created"
                
                # Save workbook
                wb.save(str(excel_path))
                
                log("INFO", f"Excel {operation} for quote {quote.id}", 
                    module="localfs_provider", tenant_id=tenant_id)
                
                return FileOperationResult(
                    success=True,
                    message=f"Quote {operation} successfully in Excel",
                    details={
                        "operation": operation,
                        "quote_id": str(quote.id),
                        "row_index": row_index or ws.max_row,
                        "excel_path": str(excel_path)
                    },
                    file_path=str(excel_path),
                    file_size=excel_path.stat().st_size,
                    modified_at=datetime.fromtimestamp(excel_path.stat().st_mtime)
                )
                
        except Timeout:
            error_msg = f"Failed to acquire file lock within {self.lock_timeout}s"
            log("ERROR", error_msg, module="localfs_provider", tenant_id=tenant_id)
            return FileOperationResult(
                success=False,
                message=error_msg,
                details={"timeout": self.lock_timeout, "lock_path": str(lock_path)}
            )
        
        except Exception as exc:
            tb = traceback.format_exc()
            log("ERROR", f"Excel update error: {exc}", module="localfs_provider", 
                tenant_id=tenant_id)
            return FileOperationResult(
                success=False,
                message=f"Excel update failed: {exc}",
                details={"error": str(exc), "traceback": tb}
            )
    
    async def read_file(self, path: str) -> bytes:
        """Read file contents as bytes."""
        resolved_path = self._resolve_path(path)
        
        if not resolved_path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        
        async with aiofiles.open(resolved_path, "rb") as f:
            return await f.read()
    
    async def write_file(self, path: str, data: bytes) -> FileOperationResult:
        """Write bytes to file."""
        resolved_path = self._resolve_path(path)
        
        try:
            # Ensure directory exists
            if self.create_dirs:
                resolved_path.parent.mkdir(parents=True, exist_ok=True)
            
            async with aiofiles.open(resolved_path, "wb") as f:
                await f.write(data)
            
            stat = resolved_path.stat()
            
            return FileOperationResult(
                success=True,
                message=f"File written successfully: {path}",
                file_path=str(resolved_path),
                file_size=stat.st_size,
                modified_at=datetime.fromtimestamp(stat.st_mtime)
            )
        
        except Exception as exc:
            log("ERROR", f"Write file error: {exc}", module="localfs_provider")
            return FileOperationResult(
                success=False,
                message=f"Failed to write file: {exc}",
                details={"error": str(exc), "traceback": traceback.format_exc()}
            )
    
    async def file_exists(self, path: str) -> bool:
        """Check if file exists."""
        try:
            resolved_path = self._resolve_path(path)
            return resolved_path.exists()
        except Exception:
            return False
    
    async def get_metadata(self, path: str) -> FileMetadata:
        """Get file metadata."""
        resolved_path = self._resolve_path(path)
        
        if not resolved_path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        
        stat = resolved_path.stat()
        
        return FileMetadata(
            path=path,
            size=stat.st_size,
            created_at=datetime.fromtimestamp(stat.st_ctime),
            modified_at=datetime.fromtimestamp(stat.st_mtime),
            exists=True,
            is_directory=resolved_path.is_dir(),
            mime_type=self._guess_mime_type(resolved_path),
            provider_metadata={
                "absolute_path": str(resolved_path),
                "permissions": oct(stat.st_mode)[-3:],
                "owner_uid": stat.st_uid,
                "owner_gid": stat.st_gid
            }
        )
    
    async def list_files(self, directory: str) -> List[FileMetadata]:
        """List files in directory."""
        resolved_dir = self._resolve_path(directory)
        
        if not resolved_dir.exists():
            raise FileNotFoundError(f"Directory not found: {directory}")
        
        if not resolved_dir.is_dir():
            raise NotADirectoryError(f"Not a directory: {directory}")
        
        files = []
        for item in resolved_dir.iterdir():
            try:
                relative_path = str(item.relative_to(self.base_path))
                metadata = await self.get_metadata(relative_path)
                files.append(metadata)
            except Exception as exc:
                log("WARNING", f"Failed to get metadata for {item}: {exc}", 
                    module="localfs_provider")
                continue
        
        return files
    
    async def delete_file(self, path: str) -> FileOperationResult:
        """Delete file."""
        resolved_path = self._resolve_path(path)
        
        try:
            if not resolved_path.exists():
                return FileOperationResult(
                    success=False,
                    message=f"File not found: {path}"
                )
            
            if resolved_path.is_dir():
                resolved_path.rmdir()
            else:
                resolved_path.unlink()
            
            return FileOperationResult(
                success=True,
                message=f"File deleted successfully: {path}",
                file_path=str(resolved_path)
            )
        
        except Exception as exc:
            log("ERROR", f"Delete file error: {exc}", module="localfs_provider")
            return FileOperationResult(
                success=False,
                message=f"Failed to delete file: {exc}",
                details={"error": str(exc), "traceback": traceback.format_exc()}
            )
    
    async def create_directory(self, path: str) -> FileOperationResult:
        """Create directory (and parents if needed)."""
        resolved_path = self._resolve_path(path)
        
        try:
            resolved_path.mkdir(parents=True, exist_ok=True)
            
            return FileOperationResult(
                success=True,
                message=f"Directory created: {path}",
                file_path=str(resolved_path)
            )
        
        except Exception as exc:
            log("ERROR", f"Create directory error: {exc}", module="localfs_provider")
            return FileOperationResult(
                success=False,
                message=f"Failed to create directory: {exc}",
                details={"error": str(exc), "traceback": traceback.format_exc()}
            )
    
    async def health_check(self) -> Dict[str, Any]:
        """
        Check provider health.
        
        Verifies:
        - Base path exists
        - Base path is accessible
        - Read permissions
        - Write permissions
        """
        checks = {
            "base_path_exists": False,
            "base_path_readable": False,
            "base_path_writable": False,
            "can_create_dirs": False,
            "disk_space_available": None
        }
        
        try:
            # Check if base path exists
            if self.base_path.exists():
                checks["base_path_exists"] = True
                
                # Check read permissions
                if os.access(self.base_path, os.R_OK):
                    checks["base_path_readable"] = True
                
                # Check write permissions
                if os.access(self.base_path, os.W_OK):
                    checks["base_path_writable"] = True
                    
                    # Try creating a test directory
                    test_dir = self.base_path / ".health_check_test"
                    try:
                        test_dir.mkdir(exist_ok=True)
                        test_dir.rmdir()
                        checks["can_create_dirs"] = True
                    except Exception:
                        pass
                
                # Check disk space
                try:
                    stat = os.statvfs(self.base_path)
                    free_space_gb = (stat.f_bavail * stat.f_frsize) / (1024**3)
                    checks["disk_space_available"] = f"{free_space_gb:.2f} GB"
                except Exception:
                    pass
            
            # Determine overall health
            healthy = (
                checks["base_path_exists"] and 
                checks["base_path_readable"] and 
                checks["base_path_writable"]
            )
            
            if healthy:
                message = "LocalFS provider healthy"
            else:
                message = "LocalFS provider unhealthy: "
                issues = []
                if not checks["base_path_exists"]:
                    issues.append("base path doesn't exist")
                if not checks["base_path_readable"]:
                    issues.append("no read access")
                if not checks["base_path_writable"]:
                    issues.append("no write access")
                message += ", ".join(issues)
            
            return {
                "healthy": healthy,
                "provider": "localfs",
                "message": message,
                "details": {
                    "base_path": str(self.base_path),
                    "checks": checks,
                    "config": {
                        "excel_filename": self.excel_filename,
                        "lock_timeout": self.lock_timeout,
                        "create_dirs": self.create_dirs
                    }
                }
            }
        
        except Exception as exc:
            return {
                "healthy": False,
                "provider": "localfs",
                "message": f"Health check failed: {exc}",
                "details": {
                    "error": str(exc),
                    "traceback": traceback.format_exc(),
                    "checks": checks
                }
            }
    
    def _guess_mime_type(self, path: Path) -> Optional[str]:
        """Guess MIME type from file extension."""
        import mimetypes
        mime_type, _ = mimetypes.guess_type(str(path))
        return mime_type
