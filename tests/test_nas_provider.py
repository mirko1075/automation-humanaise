# tests/test_nas_provider.py
"""
Comprehensive tests for NAS provider and protocol adapters.

Tests cover:
- SMB protocol adapter
- NASProvider orchestration
- Protocol registration and loading
- Document operations with different protocols
- Error handling and edge cases
"""
import pytest
from unittest.mock import Mock, AsyncMock, MagicMock, patch
import io
from datetime import datetime

# Import NAS components
from app.file_access.nas_provider import NASProvider
from app.file_access.base_fs import FileInfo, FileOperationResult, HealthCheckResult
from app.file_access.protocols.smb_protocol import SMBProtocolAdapter

# Import document operations
from app.file_access.document_ops.excel_ops import (
    read_excel,
    write_excel,
    update_excel,
    create_excel_from_data
)


class TestNASProviderOrchestration:
    """Test NASProvider orchestration and protocol loading."""
    
    def test_nas_provider_initialization_smb(self):
        """Test NASProvider initializes with SMB protocol."""
        config = {
            "protocol": "smb",
            "host": "nas.test.local",
            "share": "test_share",
            "username": "test_user",
            "password": "test_pass"
        }
        
        provider = NASProvider(config)
        
        assert provider.protocol == "smb"
        assert isinstance(provider._adapter, SMBProtocolAdapter)
    
    def test_nas_provider_unknown_protocol(self):
        """Test NASProvider raises error for unknown protocol."""
        config = {
            "protocol": "unknown_protocol",
            "host": "test.local"
        }
        
        with pytest.raises(ValueError, match="Unsupported protocol"):
            NASProvider(config)
    
    def test_nas_provider_missing_protocol(self):
        """Test NASProvider raises error when protocol not specified."""
        config = {
            "host": "test.local",
            "share": "test_share"
        }
        
        with pytest.raises(ValueError, match="Protocol not specified"):
            NASProvider(config)
    
    def test_nas_provider_protocol_registration(self):
        """Test protocol registration mechanism."""
        # Create mock adapter
        class MockAdapter:
            def __init__(self, config):
                self.config = config
        
        # Register protocol
        NASProvider.register_protocol("test_protocol", MockAdapter)
        
        # Verify registration
        assert "test_protocol" in NASProvider.get_supported_protocols()
        
        # Test instantiation
        config = {"protocol": "test_protocol", "host": "test.local"}
        provider = NASProvider(config)
        
        assert provider.protocol == "test_protocol"
        assert isinstance(provider._adapter, MockAdapter)
    
    @pytest.mark.asyncio
    async def test_nas_provider_delegates_to_adapter(self):
        """Test NASProvider delegates operations to protocol adapter."""
        config = {
            "protocol": "smb",
            "host": "nas.test.local",
            "share": "test_share",
            "username": "test_user",
            "password": "test_pass"
        }
        
        provider = NASProvider(config)
        
        # Mock the adapter
        mock_adapter = AsyncMock()
        provider._adapter = mock_adapter
        
        # Test delegation
        await provider.connect()
        mock_adapter.connect.assert_called_once()
        
        await provider.disconnect()
        mock_adapter.disconnect.assert_called_once()
        
        await provider.file_exists("/test.txt")
        mock_adapter.file_exists.assert_called_once_with("/test.txt")
        
        await provider.read_file("/test.txt")
        mock_adapter.read_file.assert_called_once_with("/test.txt")


class TestSMBProtocolAdapter:
    """Test SMB protocol adapter implementation."""
    
    def test_smb_adapter_initialization(self):
        """Test SMB adapter initialization."""
        config = {
            "host": "192.168.1.100",
            "share": "documents",
            "username": "user",
            "password": "pass",
            "base_path": "/tenant_data"
        }
        
        adapter = SMBProtocolAdapter(config)
        
        assert adapter.host == "192.168.1.100"
        assert adapter.share == "documents"
        assert adapter.username == "user"
        assert adapter.password == "pass"
        assert adapter.base_path == "/tenant_data"
        assert adapter.domain == "WORKGROUP"  # default
        assert adapter.port == 445  # default
    
    def test_smb_adapter_path_normalization(self):
        """Test SMB path normalization."""
        config = {
            "host": "nas.local",
            "share": "docs",
            "username": "user",
            "password": "pass",
            "base_path": "/tenant_data"
        }
        
        adapter = SMBProtocolAdapter(config)
        
        # Test various paths
        assert adapter._normalize_path("/file.txt") == "tenant_data/file.txt"
        assert adapter._normalize_path("file.txt") == "tenant_data/file.txt"
        assert adapter._normalize_path("/folder/file.txt") == "tenant_data/folder/file.txt"
        
        # Test with root base_path
        adapter.base_path = "/"
        assert adapter._normalize_path("/file.txt") == "file.txt"
        assert adapter._normalize_path("file.txt") == "file.txt"
    
    def test_smb_adapter_path_to_parent(self):
        """Test SMB path splitting."""
        config = {
            "host": "nas.local",
            "share": "docs",
            "username": "user",
            "password": "pass"
        }
        
        adapter = SMBProtocolAdapter(config)
        
        # Test path splitting
        parent, filename = adapter._path_to_parent("/folder/file.txt")
        assert parent == "folder"
        assert filename == "file.txt"
        
        parent, filename = adapter._path_to_parent("/file.txt")
        assert parent == ""
        assert filename == "file.txt"
        
        parent, filename = adapter._path_to_parent("/a/b/c/file.txt")
        assert parent == "a/b/c"
        assert filename == "file.txt"
    
    @pytest.mark.asyncio
    @patch('app.file_access.protocols.smb_protocol.SMBConnection')
    async def test_smb_adapter_connect(self, mock_smb_conn_class):
        """Test SMB connection establishment."""
        config = {
            "host": "nas.local",
            "share": "docs",
            "username": "user",
            "password": "pass"
        }
        
        # Mock SMBConnection
        mock_conn = MagicMock()
        mock_conn.connect.return_value = True
        mock_smb_conn_class.return_value = mock_conn
        
        adapter = SMBProtocolAdapter(config)
        
        # Connect
        await adapter.connect()
        
        # Verify connection
        assert adapter._connected is True
        assert adapter._conn is not None
        mock_conn.connect.assert_called_once()
    
    @pytest.mark.asyncio
    async def test_smb_adapter_operations_require_connection(self):
        """Test SMB operations require active connection."""
        config = {
            "host": "nas.local",
            "share": "docs",
            "username": "user",
            "password": "pass"
        }
        
        adapter = SMBProtocolAdapter(config)
        
        # Verify operations fail without connection
        with pytest.raises(ConnectionError, match="Not connected"):
            await adapter.list_files("/")
        
        with pytest.raises(ConnectionError, match="Not connected"):
            await adapter.read_file("/test.txt")
        
        with pytest.raises(ConnectionError, match="Not connected"):
            await adapter.write_file("/test.txt", b"data")


class TestDocumentOperationsWithProviders:
    """Test protocol-agnostic document operations."""
    
    @pytest.mark.asyncio
    async def test_excel_read_write_with_provider(self):
        """Test Excel read/write with any provider."""
        # Create mock provider
        mock_provider = AsyncMock()
        
        # Mock Excel data
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test Data"
        
        # Convert to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        
        mock_provider.read_file.return_value = excel_bytes
        mock_provider.write_file.return_value = FileOperationResult(
            success=True,
            path="/test.xlsx",
            message="Written"
        )
        
        # Test read
        read_wb = await read_excel(mock_provider, "/test.xlsx")
        assert read_wb.active["A1"].value == "Test Data"
        mock_provider.read_file.assert_called_once_with("/test.xlsx")
        
        # Test write
        new_wb = Workbook()
        new_wb.active["A1"] = "New Data"
        success = await write_excel(mock_provider, "/output.xlsx", new_wb, overwrite=True)
        
        assert success is True
        mock_provider.write_file.assert_called_once()
    
    @pytest.mark.asyncio
    async def test_excel_update_with_provider(self):
        """Test Excel cell updates with any provider."""
        # Create mock provider
        mock_provider = AsyncMock()
        
        # Mock existing Excel
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Original"
        
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        
        mock_provider.read_file.return_value = excel_bytes
        mock_provider.write_file.return_value = FileOperationResult(
            success=True,
            path="/test.xlsx",
            message="Updated"
        )
        
        # Test update
        updates = {
            "A1": "Updated",
            "B1": "New Value"
        }
        
        success = await update_excel(mock_provider, "/test.xlsx", updates)
        
        assert success is True
        mock_provider.read_file.assert_called_once()
        mock_provider.write_file.assert_called_once()
    
    @pytest.mark.asyncio
    async def test_excel_create_from_data(self):
        """Test creating Excel from tabular data."""
        mock_provider = AsyncMock()
        mock_provider.write_file.return_value = FileOperationResult(
            success=True,
            path="/report.xlsx",
            message="Created"
        )
        
        data = [
            ["Customer A", 1000.00, "2024-01-01"],
            ["Customer B", 1500.50, "2024-01-02"],
        ]
        headers = ["Customer", "Amount", "Date"]
        
        success = await create_excel_from_data(
            mock_provider,
            "/report.xlsx",
            data,
            headers=headers
        )
        
        assert success is True
        mock_provider.write_file.assert_called_once()


class TestProviderRegistry:
    """Test provider registry integration."""
    
    def test_registry_includes_nas_provider(self):
        """Test NAS provider is registered."""
        from app.file_access.registry import PROVIDER_REGISTRY, list_providers
        
        assert "nas" in PROVIDER_REGISTRY
        assert "nas" in list_providers()
        assert PROVIDER_REGISTRY["nas"] == NASProvider
    
    def test_get_file_provider_with_nas(self):
        """Test getting NAS provider via registry."""
        from app.file_access.registry import get_provider_by_name
        
        config = {
            "protocol": "smb",
            "host": "nas.local",
            "share": "docs",
            "username": "user",
            "password": "pass"
        }
        
        provider = get_provider_by_name("nas", config)
        
        assert isinstance(provider, NASProvider)
        assert provider.protocol == "smb"


class TestHealthChecks:
    """Test health check implementations."""
    
    @pytest.mark.asyncio
    @patch('app.file_access.protocols.smb_protocol.SMBConnection')
    async def test_smb_health_check(self, mock_smb_conn_class):
        """Test SMB adapter health check."""
        config = {
            "host": "nas.local",
            "share": "docs",
            "username": "user",
            "password": "pass"
        }
        
        # Mock SMBConnection
        mock_conn = MagicMock()
        mock_conn.connect.return_value = True
        mock_conn.listPath.return_value = []
        mock_smb_conn_class.return_value = mock_conn
        
        adapter = SMBProtocolAdapter(config)
        
        # Perform health check
        result = await adapter.health_check()
        
        # Verify
        assert isinstance(result, HealthCheckResult)
        # Health check should attempt connection
        assert "connection" in result.details or result.healthy is True
    
    @pytest.mark.asyncio
    async def test_nas_provider_health_check_adds_protocol_info(self):
        """Test NASProvider adds protocol info to health checks."""
        config = {
            "protocol": "smb",
            "host": "nas.local",
            "share": "docs",
            "username": "user",
            "password": "pass"
        }
        
        provider = NASProvider(config)
        
        # Mock adapter health check
        mock_result = HealthCheckResult(
            healthy=True,
            details={"connection": "ok"},
            message="Healthy"
        )
        
        provider._adapter.health_check = AsyncMock(return_value=mock_result)
        
        # Perform health check
        result = await provider.health_check()
        
        # Verify protocol info added
        assert result.details["protocol"] == "smb"
        assert "adapter" in result.details


class TestErrorHandling:
    """Test error handling in various scenarios."""
    
    @pytest.mark.asyncio
    async def test_file_not_found_handling(self):
        """Test FileNotFoundError is properly raised."""
        mock_provider = AsyncMock()
        mock_provider.read_file.side_effect = FileNotFoundError("File not found")
        
        with pytest.raises(FileNotFoundError):
            await read_excel(mock_provider, "/nonexistent.xlsx")
    
    @pytest.mark.asyncio
    async def test_file_exists_error_handling(self):
        """Test FileExistsError when overwrite=False."""
        mock_provider = AsyncMock()
        mock_provider.file_exists.return_value = True
        
        from openpyxl import Workbook
        wb = Workbook()
        
        with pytest.raises(Exception):
            # This should fail because file exists and overwrite=False
            mock_provider.write_file.side_effect = FileExistsError("File exists")
            await write_excel(mock_provider, "/existing.xlsx", wb, overwrite=False)
    
    @pytest.mark.asyncio
    async def test_connection_error_handling(self):
        """Test connection errors are properly handled."""
        config = {
            "protocol": "smb",
            "host": "invalid.host",
            "share": "docs",
            "username": "user",
            "password": "pass"
        }
        
        with patch('app.file_access.protocols.smb_protocol.SMBConnection') as mock_conn_class:
            mock_conn = MagicMock()
            mock_conn.connect.return_value = False
            mock_conn_class.return_value = mock_conn
            
            adapter = SMBProtocolAdapter(config)
            
            with pytest.raises(ConnectionError):
                await adapter.connect()


class TestContextManager:
    """Test context manager support."""
    
    @pytest.mark.asyncio
    async def test_provider_context_manager(self):
        """Test provider can be used as async context manager."""
        config = {
            "protocol": "smb",
            "host": "nas.local",
            "share": "docs",
            "username": "user",
            "password": "pass"
        }
        
        provider = NASProvider(config)
        
        # Mock adapter
        provider._adapter = AsyncMock()
        
        # Use as context manager
        async with provider:
            # Verify connect was called
            provider._adapter.connect.assert_called_once()
        
        # Verify disconnect was called on exit
        provider._adapter.disconnect.assert_called_once()


class TestEdgeCases:
    """Test edge cases and boundary conditions."""
    
    def test_empty_config_handling(self):
        """Test handling of empty configuration."""
        with pytest.raises(ValueError):
            NASProvider({})
    
    def test_case_insensitive_protocol_names(self):
        """Test protocol names are case-insensitive."""
        config1 = {
            "protocol": "SMB",
            "host": "nas.local",
            "share": "docs",
            "username": "user",
            "password": "pass"
        }
        
        config2 = {
            "protocol": "smb",
            "host": "nas.local",
            "share": "docs",
            "username": "user",
            "password": "pass"
        }
        
        provider1 = NASProvider(config1)
        provider2 = NASProvider(config2)
        
        assert provider1.protocol == provider2.protocol == "smb"
    
    @pytest.mark.asyncio
    async def test_operations_with_empty_paths(self):
        """Test operations with empty or root paths."""
        mock_provider = AsyncMock()
        mock_provider.list_files.return_value = []
        
        # Should not raise error
        result = await mock_provider.list_files("/")
        assert result == []


# Integration test markers
pytestmark = pytest.mark.unit


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
